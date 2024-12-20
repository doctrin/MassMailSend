import sys
import os
from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from PyQt5.QtWebEngineWidgets import QWebEngineView
from openpyxl import load_workbook, Workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import ssl
import pygame  # pygame을 사용하여 소리 재생
import re  # 이메일 패턴 검증을 위한 정규식 모듈
import webbrowser  # 웹 브라우저 열기를 위한 모듈 추가

class EmailSenderApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        # EXE 내부에서 UI 파일을 로드하기 위해 경로 설정
        ui_file = self.resource_path("email_sender_add_bodyTxt_image.ui")
        uic.loadUi(ui_file, self)

        # 버튼 클릭 이벤트 연결
        self.btnUploadExcel.clicked.connect(self.load_excel_file)
        self.btnSendEmails.clicked.connect(self.send_emails)
        self.btnClearLog.clicked.connect(self.clear_log)  # 로그 초기화 버튼 클릭 이벤트
        self.btnSaveLogToExcel.clicked.connect(self.save_log_to_excel)  # 로그 저장 버튼 클릭 이벤트
        self.btnAttachImage1.clicked.connect(lambda: self.attach_image(1))  # 이미지 1 첨부 버튼 클릭 이벤트
        self.btnAttachImage2.clicked.connect(lambda: self.attach_image(2))  # 이미지 2 첨부 버튼 클릭 이벤트
        self.btnAttachImage3.clicked.connect(lambda: self.attach_image(3))  # 이미지 3 첨부 버튼 클릭 이벤트
        self.btnRemoveImage1.clicked.connect(lambda: self.remove_image(1))  # 이미지 1 삭제 버튼 클릭 이벤트
        self.btnRemoveImage2.clicked.connect(lambda: self.remove_image(2))  # 이미지 2 삭제 버튼 클릭 이벤트
        self.btnRemoveImage3.clicked.connect(lambda: self.remove_image(3))  # 이미지 3 삭제 버튼 클릭 이벤트
        self.btnPreviewEmail.clicked.connect(self.preview_email)  # 미리보기 버튼 클릭 이벤트

        # 이미지 크기 적용 버튼 연결
        self.btnApplyImage1Size.clicked.connect(lambda: self.set_image_size(1))
        self.btnApplyImage2Size.clicked.connect(lambda: self.set_image_size(2))
        self.btnApplyImage3Size.clicked.connect(lambda: self.set_image_size(3))

        # pygame 초기화
        pygame.mixer.init()

        # 첨부된 이미지 경로 및 크기
        self.attached_images = {1: None, 2: None, 3: None}
        self.image_sizes = {1: (None, None), 2: (None, None), 3: (None, None)}

    def resource_path(self, relative_path):
        """ EXE에서 리소스 경로를 찾을 수 있도록 도와주는 함수 """
        try:
            # PyInstaller로 빌드된 경우
            base_path = sys._MEIPASS
        except Exception:
            # 개발 중일 경우
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def load_excel_file(self):
        try:

            self.textEditLog.append("엑셀 파일 로드 함수 시작")
            print("엑셀 데이터 파싱 중...")

            file_name, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel Files (*.xlsx *.xls)")
            if file_name:
                self.textEditLog.append(f"선택된 파일: {file_name}")
                self.parse_excel(file_name)
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 파일을 로드하는 중 오류가 발생했습니다:\n{str(e)}")

    def parse_excel(self, file_name):
        try:
            workbook = load_workbook(file_name, data_only=True)  # data_only 옵션으로 값만 로드
            sheet = workbook.active

            # 테이블 초기화
            self.tableWidget.setRowCount(sheet.max_row - 1)  # 첫 번째 줄은 헤더로 사용되므로 한 줄 줄여서 시작
            self.tableWidget.setColumnCount(1)  # 엑셀에서 이메일 주소만 가져옴

            # 테이블에 헤더 추가
            headers = ["이메일 주소"]
            self.tableWidget.setHorizontalHeaderLabels(headers)

            # 엑셀 데이터 테이블에 삽입 (2번째 줄부터 시작)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=0):
                email_cell = row[0]  # 첫 번째 열이 이메일 주소라고 가정
                if email_cell.value:  # 빈 값이 아닐 때만 처리
                    self.tableWidget.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(str(email_cell.value)))

            self.textEditLog.append("엑셀 파일 로드 완료!")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 데이터를 파싱하는 중 오류가 발생했습니다:\n{str(e)}")

    def validate_email(self, email):
        """ 이메일 주소가 유효한지 정규식으로 검증하는 함수 """
        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        return re.match(email_pattern, email) is not None

    def attach_image(self, slot):
        """ 이미지 파일을 선택하고 경로를 저장 """
        file_name, _ = QFileDialog.getOpenFileName(self, f"이미지 {slot} 파일 선택", "","Image Files (*.png *.jpg *.jpeg *.gif)")
        if file_name:
            self.attached_images[slot] = file_name
            self.textEditLog.append(f"이미지 {slot} 첨부: {file_name}")
            print(f"DEBUG: 이미지 {slot} 경로 -> {file_name}")

    def remove_image(self, slot):
        """ 특정 이미지 첨부를 제거 """
        self.attached_images[slot] = None
        self.textEditLog.append(f"이미지 {slot} 삭제 완료!")

    def set_image_size(self, slot):
        try:
            width = int(self.findChild(QtWidgets.QLineEdit, f"lineEditImage{slot}Width").text())
            height = int(self.findChild(QtWidgets.QLineEdit, f"lineEditImage{slot}Height").text())
            self.image_sizes[slot] = (width, height)
            self.textEditLog.append(f"이미지 {slot} 크기 설정: {width}x{height}")
        except ValueError:
            QMessageBox.warning(self, "경고", f"이미지 {slot} 크기를 올바르게 입력하세요.")

    def preview_email(self):
        """ 미리보기 팝업을 웹 브라우저로 열기 """
        import shutil  # 임시 파일 복사를 위해 추가

        subject = self.lineEditEmailSubject.text()
        body = self.textEditEmailBody.toPlainText()

        # 임시 디렉토리 설정
        temp_dir = os.path.join(os.getcwd(), "temp_preview_images")
        os.makedirs(temp_dir, exist_ok=True)

        # HTML 미리보기 작성
        html_preview = f"""<html><body>
        <h1>{subject}</h1>
        <p>{body}</p>
        <br>
        """

        for slot in range(1, 4):
            if self.attached_images[slot]:
                # 이미지 절대 경로 생성
                image_path = os.path.abspath(self.attached_images[slot])
                if os.path.exists(image_path):
                    # 임시 디렉토리로 이미지 복사
                    temp_image_path = os.path.join(temp_dir, f"image_{slot}.png")
                    shutil.copy(image_path, temp_image_path)

                    # 경로를 수정하여 HTML에 추가
                    temp_image_url = temp_image_path.replace(os.sep, "/")  # os.sep 사용
                    width, height = self.image_sizes[slot]
                    size_style = f"width:{width}px; height:{height}px;" if width and height else ""
                    html_preview += f'<img src="file:///{temp_image_url}" style="{size_style}"><br><br>'
                else:
                    self.textEditLog.append(f"이미지 {slot} 경로 오류: {image_path}가 존재하지 않습니다.")

        html_preview += "</body></html>"

        # HTML 파일 저장
        debug_file_path = os.path.join(temp_dir, "preview_debug.html")
        with open(debug_file_path, "w", encoding="utf-8") as debug_file:
            debug_file.write(html_preview)
        print(f"DEBUG: 미리보기 HTML 파일 생성 완료 -> {debug_file_path}")

        # 기본 웹 브라우저에서 HTML 파일 열기
        debug_file_url = debug_file_path.replace(os.sep, "/")  # os.sep 사용
        webbrowser.open(f"file:///{debug_file_url}")
        self.textEditLog.append(f"미리보기 브라우저 열림: {debug_file_path}")

    def send_emails(self):
        sender_email = self.lineEditSenderEmail.text()  # 발송자 이메일
        sender_password = self.lineEditSenderPassword.text()  # 발송자 비밀번호
        subject = self.lineEditEmailSubject.text()  # UI에서 입력받은 메일 제목
        body = self.textEditEmailBody.toPlainText()  # UI에서 입력받은 메일 본문

        # 빈 값 체크
        if not sender_email or not sender_password or not subject or not body:
            QMessageBox.warning(self, "경고", "이메일 주소, 비밀번호, 제목, 본문을 모두 입력해주세요!")
            return

        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        all_success = True  # 모든 메일 발송 성공 여부

        success_log = []
        failure_log = []

        try:
            # Gmail SMTP 서버에 연결
            context = ssl.create_default_context()
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls(context=context)
            server.login(sender_email, sender_password)

            self.textEditLog.append("SMTP 서버에 로그인 성공!")

            # 테이블에서 데이터 가져와서 이메일 발송
            for row in range(self.tableWidget.rowCount()):
                recipient = self.tableWidget.item(row, 0).text()  # 첫 번째 열에 이메일 주소

                # 이메일 주소 검증
                if not self.validate_email(recipient):
                    failure_log.append(f"실패: {recipient} - 이메일 주소 형식 오류")
                    self.textEditLog.append(f"{recipient} - 이메일 형식 오류!")
                    all_success = False
                    continue

                msg = MIMEMultipart()
                msg["From"] = sender_email
                msg["To"] = recipient
                msg["Subject"] = subject

                # HTML 본문 추가
                html_body = f"""<html><body>{body}<br>"""
                for slot in range(1, 4):
                    if self.attached_images[slot]:
                        ###
                        #html_body += f'<img src="cid:image{slot}" style="max-width:500px;max-height:500px;"><br>'

                        width, height = self.image_sizes[slot]
                        size_style = f"width:{width}px; height:{height}px;" if width and height else ""
                        html_body += f'<img src="cid:image{slot}" style="{size_style}"><br>'


                html_body += "</body></html>"
                msg.attach(MIMEText(html_body, "html"))

                # 이미지 첨부
                for slot in range(1, 4):
                    if self.attached_images[slot]:
                        with open(self.attached_images[slot], "rb") as img:
                            mime = MIMEBase('image', 'png', filename=os.path.basename(self.attached_images[slot]))
                            mime.set_payload(img.read())
                            encoders.encode_base64(mime)
                            mime.add_header('Content-Disposition', 'attachment', filename=os.path.basename(self.attached_images[slot]))

                            mime.add_header('Content-ID', f'<image{slot}>')
                            mime.add_header('X-Attachment-Id', f'image{slot}')
                            msg.attach(mime)

                try:
                    server.sendmail(sender_email, recipient, msg.as_string())
                    success_log.append(f"성공: {recipient} - 메일 발송 완료!")
                    self.textEditLog.append(f"{recipient}에게 메일 발송 성공!")
                except Exception as e:
                    failure_log.append(f"실패: {recipient} - {str(e)}")
                    self.textEditLog.append(f"{recipient} - 메일 발송 실패! 오류: {str(e)}")
                    all_success = False

            server.quit()
            # 모든 메일 발송 성공 시 성공 사운드 재생
            if all_success:
                self.play_success_sound()

            # 성공/실패 결과 표시
            self.textEditLog.append(f"\n메일 발송 완료! 성공: {len(success_log)} / 실패: {len(failure_log)}")
            for log in success_log:
                self.textEditLog.append(log)
            for log in failure_log:
                self.textEditLog.append(log)

        except Exception as e:
            self.textEditLog.append(f"SMTP 서버 연결 오류: {str(e)}")

    def play_success_sound(self):
        """ 메일 발송 성공 사운드 재생 """
        try:
            sound_path = self.resource_path("tada.mp3")
            pygame.mixer.music.load(sound_path)
            pygame.mixer.music.play()
            #self.textEditLog.append("성공 사운드 재생 완료!")
        except Exception as e:
            self.textEditLog.append(f"사운드 재생 오류: {str(e)}")

    def play_fail_sound(self):
        """ 메일 발송 성공 사운드 재생 """
        try:
            sound_path = self.resource_path("error.mp3")
            pygame.mixer.music.load(sound_path)
            pygame.mixer.music.play()
            #self.textEditLog.append("실패 사운드 재생 완료!")
        except Exception as e:
            self.textEditLog.append(f"사운드 재생 오류: {str(e)}")

    def clear_log(self):
        """ 로그 창을 초기화하는 함수 """
        self.textEditLog.clear()

    def save_log_to_excel(self):
        log_text = self.textEditLog.toPlainText()
        if not log_text:
            QMessageBox.warning(self, "경고", "저장할 로그가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일로 저장", "", "Excel Files (*.xlsx)")

        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Email Logs"

            for idx, line in enumerate(log_text.split("\n"), start=1):
                sheet.cell(row=idx, column=1, value=line)

            workbook.save(file_path)
            QMessageBox.information(self, "성공", f"로그가 성공적으로 저장되었습니다:\n{file_path}")

def main():
    app = QtWidgets.QApplication(sys.argv)
    window = EmailSenderApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
